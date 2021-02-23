from openpyxl import Workbook
from openpyxl import load_workbook

# 创建工作簿
wb = Workbook()

# 新增sheet
wb.create_sheet("Mysheet")
wb.create_sheet("Mysheet1")
wb.create_sheet("Mysheet2")

# wb2 = load_workbook('./docs/t1.xlsx')

print(wb.sheetnames)

currentws = wb["Sheet"]

data = [
    ['name', 'age', 'score'],
    ['asher', 11, 20],
    ['tom', 14, 55],
    ['json', 17, 38]
]

for item in data:
    currentws.append(item)

# for sheetName in wb.sheetnames:
#     print(sheetName)
#     currentWS = wb[sheetName]
#     currentCell = currentWS['A1']
#     currentCell.value = '姓名'
#     for cell in currentWS.rows:
#         print(cell)

# 保存excel
wb.save('./docs/test2.xlsx')
